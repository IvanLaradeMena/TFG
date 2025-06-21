# ───────── auto_mathcad.py ─ rellenar plantilla WCA (Prime 10) ──────────
from __future__ import annotations
from pathlib import Path
from typing import Union, Optional
import logging, shutil
import comtypes.client as cc
from comtypes import COMError
from openpyxl import load_workbook

logging.basicConfig(format="%(levelname)s: %(message)s", level=logging.INFO)

_PROGIDS = ["MathcadPrime.Application",
            "MathcadPrime.ApplicationObsolete"]          # fallback 4‑5
EXCEL_NAME = "Entrada_Datos_01.xlsx"                      # nombre fijo
# ---------------------------------------------------------------------- #
def _prime_object():
    errs = []
    for pid in _PROGIDS:
        try:
            return cc.GetActiveObject(pid)
        except (COMError, OSError):
            try:
                return cc.CreateObject(pid)
            except (COMError, OSError) as e:
                errs.append(f"{pid} → 0x{(e.args[0] if e.args else 0):08X}")
    raise RuntimeError("No se pudo conectar con Mathcad Prime; ábrelo antes.\n"
                       + "\n".join(errs))

def _ws_active(app):
    ws = getattr(app, "ActiveWorksheet", None)
    if ws is None:                       # nomenclatura 4‑5
        ws = getattr(app.Worksheets, "ActiveWorksheet", None)
    if ws is None:
        raise RuntimeError("No hay ninguna hoja activa en Mathcad Prime.")
    return ws
# ---------------------------------------------------------------------- #
def _leer_variables_excel(xlsx: Path) -> dict[str, float]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Parts Value"]
    out: dict[str, float] = {}
    for var, _tipo, val in ws.iter_rows(min_row=2, values_only=True):
        if var and val is not None:
            out[str(var).strip()] = float(val)
    return out

def _colocar_excel_junto_a_worksheet(xlsx: Path, ws) -> Path:
    destino = Path(ws.FullName).with_name(EXCEL_NAME)
    if destino.resolve() != xlsx.resolve():
        try:
            shutil.copyfile(xlsx, destino)
            logging.info("Excel actualizado en: %s", destino)
        except Exception as e:
            logging.warning("No se pudo copiar Excel: %s", e)
    return destino
# ---------------------------------------------------------------------- #
def rellenar_plantilla_wca(
    xlsx: Union[str, Path],
    plantilla: Optional[str | Path] = None
) -> None:
    xlsx = Path(xlsx).resolve()
    if not xlsx.exists():
        raise FileNotFoundError(xlsx)

    app = _prime_object()
    app.Visible = True

    if plantilla:
        plantilla = str(Path(plantilla).resolve())
        logging.info("Abriendo plantilla: %s", plantilla)
        app.Open(plantilla)

    ws = _ws_active(app)

    # 1) Copia el Excel junto a la plantilla
    _colocar_excel_junto_a_worksheet(xlsx, ws)

    # 2) Variables numéricas
    num_vars = _leer_variables_excel(xlsx)
    logging.info("Variables a transferir: %d", len(num_vars))

    no_encontradas: list[str] = []
    for var, val in num_vars.items():
        try:
            ws.SetRealValue(var, float(val), "")   
        except COMError:
            no_encontradas.append(var)

    # 3) Forzar recálculo
    for fn_name in ("Synchronize", "ResumeCalculation"):
        fn = getattr(ws, fn_name, None)
        if callable(fn):
            try:
                fn()
                break
            except COMError:
                pass

    if no_encontradas:
        logging.info("Variables no encontradas en la plantilla: %s",
                     ", ".join(no_encontradas))
    else:
        logging.info("Plantilla actualizada correctamente.")
# ---------------------------------------------------------------------- #
if __name__ == "__main__":
    import argparse, sys
    ap = argparse.ArgumentParser(
        description="Rellena una plantilla WCA de Mathcad Prime "
                    "con los valores de Entrada_Datos_01.xlsx")
    ap.add_argument("excel", help=EXCEL_NAME)
    ap.add_argument("-p", "--plantilla", help="Ruta a la plantilla .mcdx")
    args = ap.parse_args()
    try:
        rellenar_plantilla_wca(args.excel, args.plantilla)
        print("✓ Plantilla actualizada.")
    except Exception as e:
        print("ERROR:", e)
        sys.exit(1)
