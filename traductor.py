# ─────────── traductor.py  (rev-24-may-2025) ───────────
from __future__ import annotations
from pathlib import Path
import csv, re, sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook

# ───────────────────────── configuración ────────────────────────────────
DEST_XLSX = "Entrada_Datos_01.xlsx"

# ────────── numérico: '4k7' → 4700.0 ────────────────────────────────────
_SUFX = {"T":1e12, "G":1e9, "MEG":1e6, "K":1e3, "":1.0,
         "M":1e-3, "U":1e-6, "µ":1e-6, "N":1e-9, "P":1e-12, "F":1e-15}

_RE_VAL   = re.compile(r"\d+(?:[.,]\d+)?[a-zµ]{0,3}", re.I)
WARNINGS: list[str] = []          # ← aquí irán todos los avisos

def s2f(txt) -> float:
    """
    Convierte una cadena tipo “4k7” a float.
    Si la conversión falla, se añade a WARNINGS y se devuelve 0.0
    para no detener el flujo.
    """
    txt_orig = txt
    txt = "" if txt is None else str(txt)
    m = _RE_VAL.search(txt.replace(" ", ""))
    if not m:
        WARNINGS.append(f"Valor no numérico ignorado: “{txt_orig}”")
        return 0.0
    try:
        num, suf = re.match(r"^(\d+(?:[.,]\d+)?)([A-Zµ]{0,2})$", m[0].upper()).groups()
        # separar posible letra de unidad (F, H, Ω…) del prefijo (p, n, u, k…)
        suf = suf.upper()
        if suf in _SUFX:
            factor = _SUFX[suf]
        elif len(suf) > 1 and suf[:-1] in _SUFX:
            # ej. “PF” → prefix=”P”
            factor = _SUFX[suf[:-1]]
        else:
            WARNINGS.append(f"Sufijo desconocido “{suf}” en “{txt_orig}”; se asume 1")
            factor = 1.0
        return float(num.replace(",", ".")) * factor
    except Exception:
        WARNINGS.append(f"No se pudo convertir “{txt_orig}” a número")
        return 0.0

# ────────── heurística de paquete ───────────────────────────────────────
_RE_SIMPLE = re.compile(r"^([RCL])\d+$", re.I)
def guess_pkg(var: str, tol: float | None = None) -> str:
    m = _RE_SIMPLE.match(var)
    if not m:
        return var
    kind = m[1].upper()
    if kind == "R":
        return "RM0805" if tol is not None and tol <= 0.01 else "P0805"
    if kind == "C":
        return "C0805"
    if kind == "L":
        return "L0805"
    return var

DEFAULT_DEVS = {
    "RM0805": (1e-2, 100e-6, 0.0, 0.0),
    "P0805":  (2e-3,  10e-6, 0.0, 0.0),
    "C0805":  (5e-2, 200e-6, 0.0, 0.0),
}

# ────────── helpers Excel ───────────────────────────────────────────────
def _sheet(wb, name, header, rows):
    if name in wb: del wb[name]
    ws = wb.create_sheet(name)
    ws.append(header)
    for r in rows:
        ws.append(list(r))

def _wb(path: Path):
    return load_workbook(path) if path.exists() else Workbook()

# ════════════════════════════════════════════════════════════════════════
#  PARSER – LTspice
# ════════════════════════════════════════════════════════════════════════
_RE_TOLPAR  = re.compile(r"^TOL([RCL]\w+)$", re.I)
_RE_MCPAR   = re.compile(r"\{mc\(([^,]+),([^}]+)\)\}", re.I)
_RE_TEMPPAR = re.compile(r"^(?:TC|TEMP)", re.I)
_RE_AGEPAR  = re.compile(r"^AGE", re.I)
_RE_RADPAR  = re.compile(r"^RAD", re.I)

def parse_ltspice(p: Path):
    vals, pkgs, v_tols = {}, {}, {}
    grp_tol, grp_temp, grp_age, grp_rad = {}, {}, {}, {}

    txt = p.read_text(encoding="latin-1", errors="ignore").splitlines()

    # 1) .param
    for ln in txt:
        ln = ln.split(";", 1)[0].strip()
        if not ln.lower().startswith(".param"): continue
        for tok in ln[6:].split():
            if "=" not in tok: continue
            k, v = tok.split("=", 1)
            key = k.strip().upper()
            val = s2f(v)
            vals[key] = val
            pkgs[key] = key
            if _RE_TOLPAR.match(key):
                grp_tol[key] = val
                v_tols[key] = (val, 0.0, 0.0, 0.0)
            elif _RE_TEMPPAR.match(key):
                grp_temp[key] = val
                v_tols[key] = (0.0, val, 0.0, 0.0)
            elif _RE_AGEPAR.match(key):
                grp_age[key] = val
                v_tols[key] = (0.0, 0.0, val, 0.0)
            elif _RE_RADPAR.match(key):
                grp_rad[key] = val
                v_tols[key] = (0.0, 0.0, 0.0, val)
            else:
                v_tols[key] = (0.0, 0.0, 0.0, 0.0)

    # 2) R/C/L con mc()
    for ln in txt:
        ln = ln.split(";", 1)[0].strip()
        if not ln or ln[0] in ".*+": continue
        toks = ln.split()
        ref = toks[0].upper()
        if ref[0] not in "RCL": continue
        token = toks[3] if len(toks) > 3 else toks[2]

        m = _RE_MCPAR.search(token)
        if m:
            first, namep = m.group(1).strip(), m.group(2).strip().upper()
            value   = s2f(first)
            tol_val = grp_tol.get(namep, 0.0)
            tmp_val = grp_temp.get(namep, 0.0)
            age_val = grp_age.get(namep, 0.0)
            rad_val = grp_rad.get(namep, 0.0)
        else:
            value   = s2f(token)
            tol_val = tmp_val = age_val = rad_val = 0.0

        vals[ref] = value
        pkgs[ref] = guess_pkg(ref, tol_val)
        v_tols[ref] = (tol_val, tmp_val, age_val, rad_val)

    return vals, pkgs, v_tols

# ════════════════════════════════════════════════════════════════════════
#  PARSER – SIMetrix / SIMPLIS
# ════════════════════════════════════════════════════════════════════════
_RE_GAUSS = re.compile(r"\{([^}]*gauss\([^}]+\)[^}]*)\}", re.I)

def parse_simetrix(p: Path):
    vals, pkgs, v_tols = {}, {}, {}
    grp_tol, grp_temp, grp_age, grp_rad = {}, {}, {}, {}

    txt = p.read_text(encoding="utf-8", errors="ignore").splitlines()

    # .param
    for ln in txt:
        ln = ln.split(";", 1)[0].strip()
        if not ln.lower().startswith(".param"): continue
        for tok in ln.split()[1:]:
            if "=" not in tok: continue
            k, v = tok.split("=", 1)
            key = k.upper()
            val = s2f(v)
            vals[key] = val
            pkgs[key] = key
            if _RE_TOLPAR.match(key):
                grp_tol[key] = val
                v_tols[key] = (val, 0.0, 0.0, 0.0)
            elif _RE_TEMPPAR.match(key):
                grp_temp[key] = val
                v_tols[key] = (0.0, val, 0.0, 0.0)
            elif _RE_AGEPAR.match(key):
                grp_age[key] = val
                v_tols[key] = (0.0, 0.0, val, 0.0)
            elif _RE_RADPAR.match(key):
                grp_rad[key] = val
                v_tols[key] = (0.0, 0.0, 0.0, val)
            else:
                v_tols[key] = (0.0, 0.0, 0.0, 0.0)

    # R/C/L con gauss()
    for ln in txt:
        ln = ln.split(";", 1)[0].strip()
        if not ln or ln[0] in ".*": continue
        t = ln.split()
        if t[0][0] not in "RCL": continue
        ref   = t[0].upper()
        token = t[3] if len(t) > 3 else t[2]

        m = _RE_GAUSS.search(token)
        if m:
            content = m.group(1)
            parts   = content.split('*', 1)
            val     = s2f(parts[0])
            tol     = 1.0
            inner   = re.search(r"gauss\(([^)]+)\)", content, re.I)
            if inner:
                for factor in inner.group(1).split('*'):
                    tol *= s2f(factor)
        else:
            val = s2f(token)
            tol = 0.0

        vals[ref] = val
        pkgs[ref] = guess_pkg(ref, tol)
        v_tols[ref] = (tol, 0.0, 0.0, 0.0)

    return vals, pkgs, v_tols

# ════════════════════════════════════════════════════════════════════════
#  PARSER – BoM (CSV / TSV / texto plano)
# ════════════════════════════════════════════════════════════════════════
def _col(fnames, *keys):
    if not fnames: return None
    for c in fnames:
        if c and any(k in c.lower() for k in keys):
            return c
    return None

def _tokenise_plain(line: str) -> list[str]:
    return [t for t in re.split(r"\t+| {2,}", line.strip()) if t]

def parse_bom(p: Path):
    raw = p.read_text(encoding="utf-8", errors="ignore").splitlines()

    hdr_idx = next((i for i,l in enumerate(raw)
                    if "ref" in l.lower() and ("value" in l.lower()
                        or "val" in l.lower() or "part" in l.lower()
                        or "component" in l.lower())), None)
    if hdr_idx is None:
        raise ValueError("BoM: cabecera Reference / Value no encontrada")

    header_line = raw[hdr_idx]
    sample      = "\n".join(raw[hdr_idx:hdr_idx+20])

    csv_ok = True
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        csv_ok = False

    vals, pkgs, v_tols = {}, {}, {}

    if csv_ok and dialect.delimiter not in " \t":
        rdr = csv.DictReader(raw[hdr_idx:], dialect=dialect, skipinitialspace=True)
        cref = _col(rdr.fieldnames, "ref", "design")
        cval = _col(rdr.fieldnames, "value", "val", "part", "component")
        ctol = _col(rdr.fieldnames, "toler", "tol")
        ctmp = _col(rdr.fieldnames, "temp", "temperature", "tc")
        ctype= _col(rdr.fieldnames, "package", "footprint", "type")

        for row in rdr:
            if not row.get(cref) or not row.get(cval) or not re.search(r"\d", row[cval]): continue
            vnom = s2f(row[cval])
            tol = tc = 0.0
            if ctol and row.get(ctol,"").strip():
                tol = s2f(row[ctol]) / 100
            if ctmp and row.get(ctmp,"").strip():
                tc  = s2f(row[ctmp])
            pkg_row = (row.get(ctype) or "").strip()

            for ref in re.split(r"[\s,]+", row[cref].strip()):
                if not ref: continue
                r = ref.upper()
                vals[r]   = vnom
                pkgs[r]   = pkg_row or guess_pkg(r, tol)
                v_tols[r] = (tol, tc, 0.0, 0.0)
    else:
        headers = _tokenise_plain(header_line.lower())
        try:
            i_ref = next(i for i,h in enumerate(headers) if "ref" in h)
            i_val = next(i for i,h in enumerate(headers) if any(k in h for k in ("value","val","part","component")))
        except StopIteration:
            raise ValueError("BoM plano: columnas Reference / Value no encontradas")

        i_tol = next((i for i,h in enumerate(headers) if "tol" in h), None)
        i_tc  = next((i for i,h in enumerate(headers) if "temp" in h or "tc" in h), None)
        i_pkg = next((i for i,h in enumerate(headers) if "package" in h or "footprint" in h or "type" in h), None)

        for ln in raw[hdr_idx+1:]:
            if not re.search(r"\d", ln): continue
            toks = _tokenise_plain(ln)
            if len(toks) <= max(i_ref, i_val): continue
            refs = toks[i_ref]
            vnom = s2f(toks[i_val])
            tol = tc = 0.0
            if i_tol is not None and i_tol < len(toks):
                tol = s2f(toks[i_tol]) / 100
            if i_tc is not None and i_tc < len(toks):
                tc  = s2f(toks[i_tc])
            pkg_row = toks[i_pkg] if i_pkg is not None and i_pkg < len(toks) else ""

            for ref in re.split(r"[\s,]+", refs.strip()):
                if not ref: continue
                r = ref.upper()
                vals[r]   = vnom
                pkgs[r]   = pkg_row or guess_pkg(r, tol)
                v_tols[r] = (tol, tc, 0.0, 0.0)

    return vals, pkgs, v_tols

# ════════════════════════════════════════════════════════════════════════
#  PARTS-DEVIATION
# ════════════════════════════════════════════════════════════════════════
def build_devs(pkgs, var_tols):
    devs = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0])
    for var, pkg in pkgs.items():
        tol, tmp, age, rad = var_tols.get(var, (0.0, 0.0, 0.0, 0.0))
        d = devs[pkg]
        d[0] = tol or d[0]
        d[1] = tmp or d[1]
        d[2] = age or d[2]
        d[3] = rad or d[3]
    for g, d in devs.items():
        dt = DEFAULT_DEVS.get(g, (0.0, 0.0, 0.0, 0.0))
        for i in range(4):
            if d[i] == 0.0:
                d[i] = dt[i]
    return devs

# ════════════════════════════════════════════════════════════════════════
#  GENERA EXCEL
# ════════════════════════════════════════════════════════════════════════
def write_xlsx(vals, pkgs, v_tols, hs, dst=DEST_XLSX):
    wb = _wb(Path(dst))

    orden = [k for k in sorted(vals) if not k.startswith("TOL")]
    _sheet(wb, "Parts Value", ["Variable", "Tipo", "Valor"],
           [(k, pkgs[k], vals[k]) for k in orden])

    devs = build_devs(pkgs, v_tols)
    _sheet(wb, "Parts Deviation",
           ["Parametro", "Tolerancia", "Temperatura", "Ageing", "Radiation"],
           [(g, *devs[g]) for g in sorted(devs)])

    _sheet(wb, "Transfer", ["H(s)"], [(hs.strip(),)])

    if "Sheet" in wb: del wb["Sheet"]
    wb.save(dst)

# ════════════════════════════════════════════════════════════════════════
#  CLI helpers               (devuelven string con avisos incluidos)
# ════════════════════════════════════════════════════════════════════════
def _resumen_ok(txt: str) -> str:
    if WARNINGS:
        return (txt + f"\n⚠ {len(WARNINGS)} advertencia(s)\n  – "
                + "\n  – ".join(WARNINGS))
    return txt

def procesar_net(path, hs=""):
    WARNINGS.clear()
    p = Path(path)
    if "simetrix" in p.suffix.lower() or p.suffix.lower() == ".sxsch":
        vals, pkgs, tols = parse_simetrix(p)
    else:
        vals, pkgs, tols = parse_ltspice(p)
    write_xlsx(vals, pkgs, tols, hs)
    return _resumen_ok(f"✔ {len(vals)} comp ({p.name})")

def procesar_bom(path, hs=""):
    WARNINGS.clear()
    vals, pkgs, tols = parse_bom(Path(path))
    write_xlsx(vals, pkgs, tols, hs)
    return _resumen_ok(f"✔ {len(vals)} filas BoM ({Path(path).name})")

def procesar_generico(path, hs=""):
    WARNINGS.clear()
    vals, pkgs, tols = {}, {}, {}
    with Path(path).open() as f:
        for ref, val, *rest in csv.reader(f):
            r = ref.strip().upper()
            try:
                v = float(val)
            except ValueError:
                WARNINGS.append(f"Valor no numérico en CSV: “{val}” (ref {r})")
                v = 0.0
            tol = 0.0
            if rest and rest[0].strip():
                try:
                    tol = float(rest[0]) / 100
                except ValueError:
                    WARNINGS.append(f"Tolerancia no numérica en CSV: “{rest[0]}” (ref {r})")
            vals[r] = v
            pkgs[r] = guess_pkg(r, tol)
            tols[r] = (tol, 0.0, 0.0, 0.0)
    write_xlsx(vals, pkgs, tols, hs)
    return _resumen_ok(f"✔ CSV {Path(path).name}")

# ═════════════════ CLI directo ──────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("uso:  python traductor.py  archivo  [\"H(s)\"]")
    f  = Path(sys.argv[1])
    hs = sys.argv[2] if len(sys.argv) > 2 else ""
    if not f.exists():
        sys.exit("Archivo no encontrado")
    try:
        if f.suffix.lower() in {".net", ".asc", ".sxsch"}:
            print(procesar_net(f, hs))
        elif f.suffix.lower() == ".bom":
            print(procesar_bom(f, hs))
        else:
            print(procesar_generico(f, hs))
    except Exception as e:
        sys.exit(f"Error: {e}")
